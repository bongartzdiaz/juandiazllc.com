"""Build DEUS-SHARED-Pricing-Workbook.xlsx — multi-tab professional pricing reference."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from pathlib import Path

OUT = Path(__file__).parent / "DEUS-SHARED-Pricing-Workbook.xlsx"

ARIAL = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F3A5F")
HEADER_FONT = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name=ARIAL, size=10)
BOLD_BODY = Font(name=ARIAL, size=10, bold=True)
TITLE_FONT = Font(name=ARIAL, size=14, bold=True, color="1F3A5F")
NOTE_FONT = Font(name=ARIAL, size=9, italic=True, color="666666")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
EUR_FMT = '"€"#,##0.00'
PCT_FMT = "0%"
INT_FMT = "#,##0"


def style_header(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def style_body(ws, start_row, end_row, n_cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            if cell.font.name != ARIAL:
                cell.font = BODY_FONT
            cell.border = BORDER


def auto_width(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def add_title(ws, title):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 22


wb = Workbook()
wb.remove(wb.active)


# ---------------------------------------------------------------------------
# Sheet 1: Price Book
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Price Book")
add_title(ws, "DEUS-SHARED Price Book")
ws["A2"] = "All amounts in EUR, excluding VAT. Annual prices billed upfront."
ws["A2"].font = NOTE_FONT

headers = [
    "Tier", "Seat Count Min", "Commitment", "Term Months",
    "Monthly Equivalent", "Discount %", "Billed Per Seat",
    "Annual Cost Per Seat", "Stripe Price ID",
]
HEADER_ROW = 4
for i, h in enumerate(headers, 1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header(ws, HEADER_ROW, len(headers))

rows = [
    ("Operator", 1, "Monthly",       1,  40.00, 0.00,  40.00,  480.00, "price_operator_monthly"),
    ("Operator", 1, "Quarterly",     3,  36.00, 0.10, 108.00,  432.00, "price_operator_quarterly"),
    ("Operator", 1, "Semi-Annual",   6,  32.80, 0.18, 196.80,  393.60, "price_operator_semiannual"),
    ("Operator", 1, "Annual",       12,  28.00, 0.30, 336.00,  336.00, "price_operator_annual"),
    ("Team",     3, "Monthly",       1,  70.00, 0.00,  70.00,  840.00, "price_team_monthly"),
    ("Team",     3, "Quarterly",     3,  63.00, 0.10, 189.00,  756.00, "price_team_quarterly"),
    ("Team",     3, "Semi-Annual",   6,  57.40, 0.18, 344.40,  688.80, "price_team_semiannual"),
    ("Team",     3, "Annual",       12,  49.00, 0.30, 588.00,  588.00, "price_team_annual"),
    ("Business", 10, "Monthly",      1, 110.00, 0.00, 110.00, 1320.00, "price_business_monthly"),
    ("Business", 10, "Quarterly",    3,  99.00, 0.10, 297.00, 1188.00, "price_business_quarterly"),
    ("Business", 10, "Semi-Annual",  6,  90.20, 0.18, 541.20, 1082.40, "price_business_semiannual"),
    ("Business", 10, "Annual",      12,  77.00, 0.30, 924.00,  924.00, "price_business_annual"),
    ("Enterprise", 25, "Custom Annual", 12, None, None, None, None, "price_enterprise_custom"),
]

for r, row in enumerate(rows, HEADER_ROW + 1):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = LEFT if c in (1, 3, 9) else (CENTER if c in (2, 4) else RIGHT)
        if c in (5, 7, 8) and val is not None:
            cell.number_format = EUR_FMT
        if c == 6 and val is not None:
            cell.number_format = PCT_FMT
        if c in (5, 7, 8) and val is None:
            cell.value = "Contact sales"
            cell.alignment = RIGHT
        if c == 6 and val is None:
            cell.value = "Custom"
            cell.alignment = RIGHT

last_row = HEADER_ROW + len(rows)
style_body(ws, HEADER_ROW + 1, last_row, len(headers))
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{last_row}"
auto_width(ws, {
    "A": 14, "B": 14, "C": 16, "D": 13,
    "E": 20, "F": 12, "G": 18, "H": 22, "I": 32,
})


# ---------------------------------------------------------------------------
# Sheet 2: TCO Calculator
# ---------------------------------------------------------------------------
ws = wb.create_sheet("TCO Calculator")
add_title(ws, "TCO Scenarios by Team Size")
ws["A2"] = "Compare monthly vs annual commitment savings across common team sizes."
ws["A2"].font = NOTE_FONT

headers = [
    "Scenario", "Team Size", "Recommended Tier", "Commitment",
    "Monthly Per Seat (EUR)", "Monthly Total (EUR)",
    "Annual Total (EUR)", "Monthly-Plan Annual (EUR)",
    "Savings vs Monthly (EUR)", "Notes",
]
HEADER_ROW = 4
for i, h in enumerate(headers, 1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header(ws, HEADER_ROW, len(headers))

monthly_lookup = {"Operator": 40.00, "Team": 70.00, "Business": 110.00}
scenarios = [
    ("Solo agent testing",                1, "Operator", "Monthly",  40.00, "Entry point. Convert to annual after 30 days of active use."),
    ("Solo agent committed",              1, "Operator", "Annual",   28.00, "Anchor offer. Saves EUR 144/year."),
    ("2-agent partnership",               2, "Operator", "Annual",   28.00, "Stay on Operator until 3rd hire."),
    ("3-agent micro-team",                3, "Team",     "Monthly",  70.00, "Trial month before annual lock-in."),
    ("3-agent micro-team committed",      3, "Team",     "Annual",   49.00, "Saves about one seat per year."),
    ("5-agent boutique",                  5, "Team",     "Annual",   49.00, "Team scales to 9 seats before Business."),
    ("8-agent brokerage",                 8, "Team",     "Annual",   49.00, "Stay on Team unless API/SSO/white-label needed."),
    ("10-agent brokerage",               10, "Business", "Monthly", 110.00, "Trial month before commit."),
    ("10-agent brokerage committed",     10, "Business", "Annual",   77.00, "Business floor. Includes white-label + CMA."),
    ("15-agent firm",                    15, "Business", "Annual",   77.00, "Sweet spot for Business tier."),
    ("25-agent firm",                    25, "Enterprise", "Annual",  None, "Volume discount + custom DPIA negotiable."),
    ("50+ multi-office",                 50, "Enterprise", "Annual",  None, "Dedicated CSM + 99.9% SLA."),
]

for i, (scn, size, tier, commit, per_seat, notes) in enumerate(scenarios, HEADER_ROW + 1):
    ws.cell(row=i, column=1, value=scn).alignment = LEFT
    ws.cell(row=i, column=2, value=size).alignment = CENTER
    ws.cell(row=i, column=3, value=tier).alignment = LEFT
    ws.cell(row=i, column=4, value=commit).alignment = CENTER

    if per_seat is None:
        ws.cell(row=i, column=5, value="Contact sales").alignment = RIGHT
        ws.cell(row=i, column=6, value="Contact sales").alignment = RIGHT
        ws.cell(row=i, column=7, value="Contact sales").alignment = RIGHT
        ws.cell(row=i, column=8, value="Contact sales").alignment = RIGHT
        ws.cell(row=i, column=9, value="Contact sales").alignment = RIGHT
    else:
        c = ws.cell(row=i, column=5, value=per_seat)
        c.number_format = EUR_FMT
        c.alignment = RIGHT

        c = ws.cell(row=i, column=6, value=f"=B{i}*E{i}")
        c.number_format = EUR_FMT
        c.alignment = RIGHT

        c = ws.cell(row=i, column=7, value=f"=F{i}*12")
        c.number_format = EUR_FMT
        c.alignment = RIGHT

        mo_price = monthly_lookup[tier]
        c = ws.cell(row=i, column=8, value=size * mo_price * 12)
        c.number_format = EUR_FMT
        c.alignment = RIGHT

        c = ws.cell(row=i, column=9, value=f"=H{i}-G{i}")
        c.number_format = EUR_FMT
        c.alignment = RIGHT

    ws.cell(row=i, column=10, value=notes).alignment = LEFT

last_row = HEADER_ROW + len(scenarios)
style_body(ws, HEADER_ROW + 1, last_row, len(headers))
ws.freeze_panes = "B5"
ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{last_row}"

# Conditional formatting on Savings column
ws.conditional_formatting.add(
    f"I{HEADER_ROW + 1}:I{last_row}",
    ColorScaleRule(
        start_type="min", start_color="FFFFFF",
        mid_type="percentile", mid_value=50, mid_color="C6EFCE",
        end_type="max", end_color="3CB371",
    ),
)

auto_width(ws, {
    "A": 32, "B": 11, "C": 18, "D": 14,
    "E": 22, "F": 22, "G": 22, "H": 26, "I": 26, "J": 50,
})


# ---------------------------------------------------------------------------
# Sheet 3: Tier Features
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Tier Features")
add_title(ws, "Feature Matrix by Tier")
ws["A2"] = "Yes = included; No = not available; Basic/Full = partial vs complete; Custom = negotiated per contract."
ws["A2"].font = NOTE_FONT

headers = ["Feature", "Operator", "Team", "Business", "Enterprise"]
HEADER_ROW = 4
for i, h in enumerate(headers, 1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header(ws, HEADER_ROW, len(headers))

features = [
    ("Contacts, Leads, Deals, Pipeline",     "Yes",   "Yes",   "Yes",   "Yes"),
    ("Action plans and drip campaigns",      "Yes",   "Yes",   "Yes",   "Yes"),
    ("Email and SMS",                        "Yes",   "Yes",   "Yes",   "Yes"),
    ("Client portal",                        "Basic", "Full",  "Full + branding", "Full + branding"),
    ("Mobile app (iOS and Android)",         "Yes",   "Yes",   "Yes",   "Yes"),
    ("Four-locale UI (EN/NL/DE/ES)",         "Yes",   "Yes",   "Yes",   "Yes"),
    ("Multi-user and roles",                 "No",    "Yes",   "Yes",   "Yes"),
    ("Lead routing",                         "No",    "Yes",   "Yes",   "Yes"),
    ("Lead scoring",                         "No",    "Yes",   "Yes",   "Yes"),
    ("Team analytics",                       "No",    "Yes",   "Yes",   "Yes"),
    ("Audit log",                            "No",    "Yes",   "Yes",   "Yes"),
    ("Custom fields",                        "No",    "Yes",   "Yes",   "Yes"),
    ("White-label and custom domain",        "No",    "No",    "Yes",   "Yes"),
    ("API access",                           "No",    "No",    "Yes",   "Yes"),
    ("SSO (SAML)",                           "No",    "No",    "Yes",   "Yes"),
    ("CMA tool",                             "No",    "No",    "Yes",   "Yes"),
    ("Market analytics dashboard",           "No",    "No",    "Yes",   "Yes"),
    ("Advanced reporting",                   "No",    "No",    "Yes",   "Yes"),
    ("Storage per seat",                     "5 GB",  "25 GB", "100 GB", "Unlimited"),
    ("Support SLA",                          "Email best-effort", "Email 24h", "Email 4h + chat", "Dedicated CSM + 99.9% SLA"),
    ("Backup-restore drill",                 "No",    "Annual", "Quarterly", "Quarterly + counsel-reviewed"),
    ("GDPR DPA",                             "Self-serve", "Self-serve", "Counter-signed", "Counter-signed + custom RoPA"),
]

for r, row in enumerate(features, HEADER_ROW + 1):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = LEFT if c == 1 else CENTER

last_row = HEADER_ROW + len(features)
style_body(ws, HEADER_ROW + 1, last_row, len(headers))
ws.freeze_panes = "B5"
auto_width(ws, {"A": 38, "B": 22, "C": 22, "D": 30, "E": 32})


# ---------------------------------------------------------------------------
# Sheet 4: Add-ons
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Add-ons")
add_title(ws, "Add-on Pricing")
ws["A2"] = "Available across all tiers. Usage-based add-ons are billed monthly in arrears."
ws["A2"].font = NOTE_FONT

headers = ["Add-on", "Unit Price (EUR)", "Unit", "Free Quota", "Notes"]
HEADER_ROW = 4
for i, h in enumerate(headers, 1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header(ws, HEADER_ROW, len(headers))

addons = [
    ("Extra storage",          0.10,    "Per GB per month",       "Tier-included quota", "Auto-billed when tier quota exceeded."),
    ("Extra phone number",     5.00,    "Per number per month",   "0",                   "Twilio-backed. For SMS and call routing."),
    ("SMS bundle",             0.05,    "Per message",            "0",                   "EU pricing average. Varies by country."),
    ("Email overage",          0.001,   "Per email",              "10000 per seat/mo",   "First 10k emails per seat are included."),
    ("Implementation pack",    495.00,  "One-time",               "n/a",                 "Data import + 1h setup call + 30 days email support."),
    ("White-glove onboarding", 2500.00, "One-time",               "n/a",                 "Required on Business. Included in Enterprise."),
    ("DPIA / RoPA review",     750.00,  "One-time",               "n/a",                 "Recommended for healthcare and public-sector buyers."),
]

for r, row in enumerate(addons, HEADER_ROW + 1):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        if c == 2:
            cell.number_format = EUR_FMT
            cell.alignment = RIGHT
        elif c == 1:
            cell.alignment = LEFT
        else:
            cell.alignment = LEFT

last_row = HEADER_ROW + len(addons)
style_body(ws, HEADER_ROW + 1, last_row, len(headers))
ws.freeze_panes = "A5"
auto_width(ws, {"A": 28, "B": 18, "C": 24, "D": 22, "E": 58})


# ---------------------------------------------------------------------------
# Sheet 5: Discount Math
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Discount Math")
add_title(ws, "Annual Savings vs Monthly")
ws["A2"] = "Savings calculations at minimum seat counts for each tier."
ws["A2"].font = NOTE_FONT

headers = [
    "Tier", "Monthly Cost Per Seat (EUR)", "Annual Cost Per Seat (EUR)",
    "Annual Savings Per Seat (EUR)", "Annual Savings %",
    "Min Seats", "Annual Savings at Min Seats (EUR)",
]
HEADER_ROW = 4
for i, h in enumerate(headers, 1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header(ws, HEADER_ROW, len(headers))

discount_rows = [
    ("Operator", 40.00, 28.00,  1),
    ("Team",     70.00, 49.00,  3),
    ("Business", 110.00, 77.00, 10),
]

for i, (tier, mo, ann, min_seats) in enumerate(discount_rows, HEADER_ROW + 1):
    ws.cell(row=i, column=1, value=tier).alignment = LEFT
    c = ws.cell(row=i, column=2, value=mo); c.number_format = EUR_FMT; c.alignment = RIGHT
    c = ws.cell(row=i, column=3, value=ann); c.number_format = EUR_FMT; c.alignment = RIGHT
    c = ws.cell(row=i, column=4, value=f"=(B{i}-C{i})*12"); c.number_format = EUR_FMT; c.alignment = RIGHT
    c = ws.cell(row=i, column=5, value=f"=(B{i}-C{i})/B{i}"); c.number_format = PCT_FMT; c.alignment = RIGHT
    c = ws.cell(row=i, column=6, value=min_seats); c.alignment = CENTER
    c = ws.cell(row=i, column=7, value=f"=D{i}*F{i}"); c.number_format = EUR_FMT; c.alignment = RIGHT

last_row = HEADER_ROW + len(discount_rows)
style_body(ws, HEADER_ROW + 1, last_row, len(headers))
auto_width(ws, {"A": 14, "B": 28, "C": 28, "D": 30, "E": 20, "F": 12, "G": 30})


# ---------------------------------------------------------------------------
# Sheet 6: Commitment Curve
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Commitment Curve")
add_title(ws, "Commitment Discount Curve")
ws["A2"] = "Aggressive ladder: 0 / 10 / 18 / 30 percent. Annual is the anchor offer."
ws["A2"].font = NOTE_FONT

headers = [
    "Commitment", "Term Months", "Discount %",
    "Operator (EUR/seat/mo)", "Team (EUR/seat/mo)", "Business (EUR/seat/mo)",
]
HEADER_ROW = 4
for i, h in enumerate(headers, 1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header(ws, HEADER_ROW, len(headers))

curve_rows = [
    ("Monthly",      1,  0.00, 40.00, 70.00, 110.00),
    ("Quarterly",    3,  0.10, 36.00, 63.00,  99.00),
    ("Semi-Annual",  6,  0.18, 32.80, 57.40,  90.20),
    ("Annual",      12,  0.30, 28.00, 49.00,  77.00),
]

for i, (commit, term, disc, op, tm, bz) in enumerate(curve_rows, HEADER_ROW + 1):
    ws.cell(row=i, column=1, value=commit).alignment = LEFT
    ws.cell(row=i, column=2, value=term).alignment = CENTER
    c = ws.cell(row=i, column=3, value=disc); c.number_format = PCT_FMT; c.alignment = RIGHT
    c = ws.cell(row=i, column=4, value=op); c.number_format = EUR_FMT; c.alignment = RIGHT
    c = ws.cell(row=i, column=5, value=tm); c.number_format = EUR_FMT; c.alignment = RIGHT
    c = ws.cell(row=i, column=6, value=bz); c.number_format = EUR_FMT; c.alignment = RIGHT

last_row = HEADER_ROW + len(curve_rows)
style_body(ws, HEADER_ROW + 1, last_row, len(headers))
# highlight annual row
for c in range(1, len(headers) + 1):
    ws.cell(row=last_row, column=c).fill = PatternFill("solid", start_color="FFF2CC")
    ws.cell(row=last_row, column=c).font = BOLD_BODY

auto_width(ws, {"A": 16, "B": 14, "C": 14, "D": 26, "E": 26, "F": 26})


# ---------------------------------------------------------------------------
# Sheet 7: Coupons
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Coupons")
add_title(ws, "Active Coupons")
ws["A2"] = "Coupons are tied to email-sequence triggers and single-use per customer."
ws["A2"].font = NOTE_FONT

headers = ["Code", "Description", "Discount %", "Applies To", "Validity (Days)", "Max Redemptions Per Customer"]
HEADER_ROW = 4
for i, h in enumerate(headers, 1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header(ws, HEADER_ROW, len(headers))

coupon_rows = [
    ("WELCOMEBACK10", "Trial winback offer (email-winback sequence)", 0.10, "Quarterly plan, first term only", 25, 1),
]

for i, row in enumerate(coupon_rows, HEADER_ROW + 1):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=c, value=val)
        if c == 3:
            cell.number_format = PCT_FMT
            cell.alignment = RIGHT
        elif c in (5, 6):
            cell.alignment = CENTER
        else:
            cell.alignment = LEFT

last_row = HEADER_ROW + len(coupon_rows)
style_body(ws, HEADER_ROW + 1, last_row, len(headers))
auto_width(ws, {"A": 18, "B": 42, "C": 12, "D": 32, "E": 16, "F": 28})


# ---------------------------------------------------------------------------
# Sheet 8: Stripe Mapping
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Stripe Mapping")
add_title(ws, "Stripe Product / Price Mapping")
ws["A2"] = "Operator action: create one Stripe Product per tier and one Price per commitment. Map env vars to the IDs below."
ws["A2"].font = NOTE_FONT

headers = [
    "Stripe Price ID", "Tier", "Commitment", "Amount (EUR)",
    "Recurring Interval", "Trial Period (Days)", "Tax Behavior", "Notes",
]
HEADER_ROW = 4
for i, h in enumerate(headers, 1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header(ws, HEADER_ROW, len(headers))

stripe_rows = [
    ("price_operator_monthly",      "Operator", "Monthly",      40.00, "month",  14, "Exclusive", "Self-serve trial. No card required."),
    ("price_operator_quarterly",    "Operator", "Quarterly",   108.00, "3 months",  14, "Exclusive", "Billed every 3 months."),
    ("price_operator_semiannual",   "Operator", "Semi-Annual", 196.80, "6 months",  14, "Exclusive", "Billed every 6 months."),
    ("price_operator_annual",       "Operator", "Annual",      336.00, "year",   14, "Exclusive", "Anchor offer. 30% discount."),
    ("price_team_monthly",          "Team",     "Monthly",      70.00, "month",  14, "Exclusive", "Min 3 seats enforced at checkout."),
    ("price_team_quarterly",        "Team",     "Quarterly",   189.00, "3 months",  14, "Exclusive", "Billed every 3 months."),
    ("price_team_semiannual",       "Team",     "Semi-Annual", 344.40, "6 months",  14, "Exclusive", "Billed every 6 months."),
    ("price_team_annual",           "Team",     "Annual",      588.00, "year",   14, "Exclusive", "Min 3 seats. 30% discount."),
    ("price_business_monthly",      "Business", "Monthly",     110.00, "month",   0, "Exclusive", "Sales-assisted only. Min 10 seats."),
    ("price_business_quarterly",    "Business", "Quarterly",   297.00, "3 months",  0, "Exclusive", "Sales-assisted only."),
    ("price_business_semiannual",   "Business", "Semi-Annual", 541.20, "6 months",  0, "Exclusive", "Sales-assisted only."),
    ("price_business_annual",       "Business", "Annual",      924.00, "year",   0, "Exclusive", "Sales-assisted only. White-label + SSO."),
    ("price_enterprise_custom",     "Enterprise", "Custom",     None,  "year",   0, "Exclusive", "Quoted per opportunity. Custom DPIA + integrations."),
]

for i, (pid, tier, commit, amount, interval, trial, tax, notes) in enumerate(stripe_rows, HEADER_ROW + 1):
    ws.cell(row=i, column=1, value=pid).alignment = LEFT
    ws.cell(row=i, column=2, value=tier).alignment = LEFT
    ws.cell(row=i, column=3, value=commit).alignment = CENTER
    if amount is None:
        ws.cell(row=i, column=4, value="Custom").alignment = RIGHT
    else:
        c = ws.cell(row=i, column=4, value=amount); c.number_format = EUR_FMT; c.alignment = RIGHT
    ws.cell(row=i, column=5, value=interval).alignment = CENTER
    ws.cell(row=i, column=6, value=trial).alignment = CENTER
    ws.cell(row=i, column=7, value=tax).alignment = CENTER
    ws.cell(row=i, column=8, value=notes).alignment = LEFT

last_row = HEADER_ROW + len(stripe_rows)
style_body(ws, HEADER_ROW + 1, last_row, len(headers))
ws.freeze_panes = "B5"
ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{last_row}"
auto_width(ws, {
    "A": 32, "B": 14, "C": 14, "D": 16,
    "E": 14, "F": 16, "G": 14, "H": 52,
})


# ---------------------------------------------------------------------------
# Sheet 9: Terms
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Terms")
add_title(ws, "Commercial Terms")
ws["A2"] = "Sales-team reference. All terms apply to standard contracts unless overridden in Enterprise MSA."
ws["A2"].font = NOTE_FONT

headers = ["Topic", "Term"]
HEADER_ROW = 4
for i, h in enumerate(headers, 1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header(ws, HEADER_ROW, len(headers))

terms_rows = [
    ("Refund Policy (Monthly)",    "No refund. Service paid through end of current month."),
    ("Refund Policy (Quarterly)",  "No refund. Committed for 3 months."),
    ("Refund Policy (Semi-Annual)","Pro-rata refund if SLA breached 3 or more months in a row."),
    ("Refund Policy (Annual)",     "Pro-rata refund on SLA breach, OR full refund within first 14 days for B2C buyers (EU cooling-off)."),
    ("SLA",                        "99.9% uptime guarantee on Business and Enterprise tiers."),
    ("VAT Handling",               "Stripe Tax automatically applies EU reverse-charge. B2B with valid VAT-ID exempt at checkout."),
    ("Currency",                   "EUR base. Optional USD/GBP via Stripe Adaptive Pricing (post-EU launch)."),
    ("Trial Length",               "14 days. No credit card required for Operator and Team. Business trial requires sales call."),
    ("Cancellation",               "End of current term. No auto-renewal if opted out before term-end."),
    ("Data Export",                "JSON and CSV export available for 90 days after cancellation. Earlier deletion on written request."),
    ("Grandfathered Pricing",      "Annual buyers locked at signup rate. If list prices increase, grandfathered rate continues on renewal."),
    ("B2B vs B2C",                 "EU cooling-off applies to B2C natural persons. B2B legal entities waive via welcome wizard checkbox at signup."),
    ("GDPR DPA",                   "Self-serve download on signup (Operator, Team). Counter-signed on Business+. Custom RoPA on Enterprise."),
    ("Data Residency",             "EU-hosted (Frankfurt and Amsterdam regions). No US data residency option."),
    ("Sub-processors",             "Published list with notification 30 days prior to any new sub-processor."),
    ("Acceptable Use",             "Standard SaaS AUP. No automated mass-data scraping or competitor reverse-engineering."),
]

for r, row in enumerate(terms_rows, HEADER_ROW + 1):
    cell_a = ws.cell(row=r, column=1, value=row[0])
    cell_a.alignment = LEFT
    cell_a.font = BOLD_BODY
    cell_b = ws.cell(row=r, column=2, value=row[1])
    cell_b.alignment = LEFT

last_row = HEADER_ROW + len(terms_rows)
style_body(ws, HEADER_ROW + 1, last_row, len(headers))
ws.freeze_panes = "A5"
auto_width(ws, {"A": 32, "B": 90})


# ---------------------------------------------------------------------------
# Sheet 0: Cover / TOC (insert first)
# ---------------------------------------------------------------------------
cover = wb.create_sheet("Cover", 0)
cover["A1"] = "DEUS-SHARED Pricing Workbook"
cover["A1"].font = Font(name=ARIAL, size=20, bold=True, color="1F3A5F")
cover["A2"] = "Real Estate CRM and Dashboard"
cover["A2"].font = Font(name=ARIAL, size=12, italic=True, color="666666")
cover["A4"] = "Version: 1.0"
cover["A5"] = "Last updated: 2026-05-14"
cover["A6"] = "Currency: EUR (excluding VAT)"
cover["A7"] = "Owner: juandiazllc.com"
for row in range(4, 8):
    cover.cell(row=row, column=1).font = BODY_FONT

cover["A9"] = "Workbook Contents"
cover["A9"].font = Font(name=ARIAL, size=12, bold=True, color="1F3A5F")

toc = [
    ("Price Book",        "Master pricing matrix: 13 tier x commitment rows."),
    ("TCO Calculator",    "Twelve team-size scenarios with savings calculations."),
    ("Tier Features",     "22-feature comparison across all four tiers."),
    ("Add-ons",           "Seven additional product offerings and usage fees."),
    ("Discount Math",     "Annual savings calculations at minimum seat counts."),
    ("Commitment Curve",  "Discount ladder: 0 / 10 / 18 / 30 percent."),
    ("Coupons",           "Active discount codes and validity windows."),
    ("Stripe Mapping",    "Price IDs, intervals, trial periods for Stripe setup."),
    ("Terms",             "Commercial terms: refunds, SLA, VAT, cancellation."),
]

cover.cell(row=10, column=1, value="Tab").font = HEADER_FONT
cover.cell(row=10, column=1).fill = HEADER_FILL
cover.cell(row=10, column=1).alignment = CENTER
cover.cell(row=10, column=1).border = BORDER
cover.cell(row=10, column=2, value="Description").font = HEADER_FONT
cover.cell(row=10, column=2).fill = HEADER_FILL
cover.cell(row=10, column=2).alignment = CENTER
cover.cell(row=10, column=2).border = BORDER

for i, (tab, desc) in enumerate(toc, 11):
    a = cover.cell(row=i, column=1, value=tab)
    a.font = BOLD_BODY
    a.alignment = LEFT
    a.border = BORDER
    b = cover.cell(row=i, column=2, value=desc)
    b.font = BODY_FONT
    b.alignment = LEFT
    b.border = BORDER

cover.column_dimensions["A"].width = 24
cover.column_dimensions["B"].width = 70

last_toc_row = 10 + len(toc)
cover.cell(row=last_toc_row + 2, column=1, value="Notes").font = Font(name=ARIAL, size=11, bold=True, color="1F3A5F")
cover.cell(row=last_toc_row + 3, column=1, value="All amounts in EUR excluding VAT.").font = NOTE_FONT
cover.cell(row=last_toc_row + 4, column=1, value="Annual prices are billed upfront for the full 12-month term.").font = NOTE_FONT
cover.cell(row=last_toc_row + 5, column=1, value="Stripe Price IDs are placeholders; replace with live IDs after creation.").font = NOTE_FONT
cover.cell(row=last_toc_row + 6, column=1, value="This workbook is the single source of truth for pricing; the JSON config file mirrors it.").font = NOTE_FONT


wb.save(OUT)
print(f"Workbook saved: {OUT}")
print(f"Sheets: {wb.sheetnames}")
